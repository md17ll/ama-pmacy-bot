from __future__ import annotations

import asyncio
from datetime import UTC, datetime
from io import BytesIO
from types import SimpleNamespace
from zoneinfo import ZoneInfo

from aiogram.enums import ButtonStyle
from openpyxl import load_workbook

from app import callbacks as cb, texts
from app.middlewares import AntiSpamMiddleware
from app.public_keyboards import DEVELOPER_URL, user_home, user_results
from app.services.excel import build_missing_pharmacies_template


def _all_buttons(markup):
    return [button for row in markup.inline_keyboard for button in row]


def test_home_button_colors_and_developer_link(monkeypatch) -> None:
    monkeypatch.delenv("PREMIUM_EMOJI_ID", raising=False)
    markup = user_home()
    buttons = {button.text: button for button in _all_buttons(markup)}

    assert buttons["📅 صيدليات اليوم"].style == ButtonStyle.SUCCESS
    assert buttons["⏭ صيدليات غداً"].style == ButtonStyle.SUCCESS
    assert buttons["🔄 تحديث الوقت"].style == ButtonStyle.DANGER
    assert buttons["👨‍💻 مطوّر البوت"].style == ButtonStyle.SUCCESS
    assert buttons["👨‍💻 مطوّر البوت"].url == DEVELOPER_URL


def test_home_supports_premium_custom_emoji(monkeypatch) -> None:
    monkeypatch.setenv("PREMIUM_EMOJI_ID", "5368324170671202286")
    markup = user_home()
    developer = next(button for button in _all_buttons(markup) if button.url == DEVELOPER_URL)

    assert developer.text == "مطوّر البوت"
    assert developer.icon_custom_emoji_id == "5368324170671202286"
    assert developer.style == ButtonStyle.SUCCESS


def test_results_show_pharmacies_refresh_and_blue_back() -> None:
    pharmacy_one = SimpleNamespace(id=1, name="صيدلية الشفاء")
    pharmacy_two = SimpleNamespace(id=2, name="صيدلية الأمل")
    shifts = [
        SimpleNamespace(pharmacy=pharmacy_one),
        SimpleNamespace(pharmacy=pharmacy_one),
        SimpleNamespace(pharmacy=pharmacy_two),
    ]

    markup = user_results(shifts, refresh_callback=cb.USER_TOMORROW)
    buttons = _all_buttons(markup)

    assert [button.text for button in buttons] == [
        "💊 صيدلية الشفاء",
        "💊 صيدلية الأمل",
        "🔄 تحديث الوقت",
        "⬅️ رجوع",
    ]
    assert buttons[-2].callback_data == cb.USER_TOMORROW
    assert buttons[-2].style == ButtonStyle.DANGER
    assert buttons[-1].callback_data == cb.USER_HOME
    assert buttons[-1].style == ButtonStyle.PRIMARY


def test_results_without_shifts_have_refresh_and_back() -> None:
    markup = user_results([], refresh_callback=cb.USER_NOW)
    buttons = _all_buttons(markup)

    assert [button.text for button in buttons] == ["🔄 تحديث الوقت", "⬅️ رجوع"]
    assert buttons[0].callback_data == cb.USER_NOW
    assert buttons[1].callback_data == cb.USER_HOME


def test_public_search_does_not_show_admin_status() -> None:
    pharmacy = SimpleNamespace(
        name="صيدلية الشفاء",
        address="شارع البلدية",
        status="active",
    )
    result = texts.pharmacy_result_text(
        pharmacy,
        None,
        datetime.now(UTC),
        ZoneInfo("Asia/Damascus"),
    )

    assert "الحالة" not in result
    assert "فعالة" not in result


def test_missing_pharmacies_template_is_prefilled() -> None:
    data = build_missing_pharmacies_template(
        ["صيدلية الأمل", "صيدلية الشفاء", "صيدلية الأمل"],
        batch_id=7,
    )
    workbook = load_workbook(BytesIO(data), data_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))

    assert rows[0][:2] == ("اسم الصيدلية", "العنوان")
    assert [row[0] for row in rows[1:]] == ["صيدلية الأمل", "صيدلية الشفاء"]
    assert all(row[1] in (None, "") for row in rows[1:])


def test_antispam_cooldown_and_release() -> None:
    async def scenario() -> None:
        middleware = AntiSpamMiddleware(callback_cooldown=0.65)

        assert await middleware._try_enter(100, "callback", now=10.0)
        assert not await middleware._try_enter(100, "callback", now=10.1)
        await middleware._leave(100, "callback")
        assert not await middleware._try_enter(100, "callback", now=10.2)
        assert await middleware._try_enter(100, "callback", now=10.8)
        await middleware._leave(100, "callback")

    asyncio.run(scenario())
