from __future__ import annotations

from datetime import date, datetime, time
from io import BytesIO
from pathlib import PurePosixPath
from typing import Any, Iterable
from zipfile import BadZipFile, ZipFile, is_zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.utils import ParsedShift, parse_date_value, parse_time_value


SHIFT_HEADERS = {
    "pharmacy": {"اسم الصيدلية", "الصيدلية", "pharmacy", "pharmacy name"},
    "date": {"التاريخ", "تاريخ", "date", "duty date"},
    "start": {"وقت البداية", "بداية المناوبة", "من", "start", "start time"},
    "end": {"وقت النهاية", "نهاية المناوبة", "إلى", "الى", "end", "end time"},
}

PHARMACY_HEADERS = {
    "name": {"اسم الصيدلية", "الصيدلية", "name", "pharmacy"},
    "address": {"العنوان", "عنوان", "address"},
    "aliases": {"الأسماء البديلة", "اسماء بديلة", "aliases"},
    "status": {"الحالة", "status"},
    "notes": {"ملاحظات", "notes"},
}

MAX_XLSX_BYTES = 10 * 1024 * 1024
MAX_XLSX_UNCOMPRESSED_BYTES = 64 * 1024 * 1024
MAX_XLSX_ENTRIES = 512
MAX_WORKBOOK_ROWS = 5000
MAX_WORKBOOK_COLUMNS = 32

_STATUS_ALIASES = {
    "active": "active",
    "فعالة": "active",
    "فعال": "active",
    "temporarily_closed": "temporarily_closed",
    "مغلقة مؤقتاً": "temporarily_closed",
    "مغلقة مؤقتا": "temporarily_closed",
    "inactive": "inactive",
    "متوقفة": "inactive",
    "متوقف": "inactive",
}


def _validate_xlsx_archive(data: bytes) -> None:
    if len(data) > MAX_XLSX_BYTES:
        raise ValueError("حجم ملف Excel أكبر من الحد المسموح")
    if not is_zipfile(BytesIO(data)):
        raise ValueError("الملف ليس ملف xlsx صالحاً")
    try:
        with ZipFile(BytesIO(data)) as archive:
            entries = archive.infolist()
            if len(entries) > MAX_XLSX_ENTRIES:
                raise ValueError("ملف Excel يحتوي على عدد ملفات داخلي كبير جداً")
            total_size = 0
            for entry in entries:
                path = PurePosixPath(entry.filename)
                if path.is_absolute() or ".." in path.parts:
                    raise ValueError("ملف Excel يحتوي على مسار داخلي غير آمن")
                total_size += entry.file_size
                if total_size > MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise ValueError("محتوى ملف Excel بعد فك الضغط أكبر من الحد المسموح")
    except BadZipFile as exc:
        raise ValueError("تعذر فتح بنية ملف Excel") from exc


def _validate_sheet_size(sheet) -> None:
    if sheet.max_column and sheet.max_column > MAX_WORKBOOK_COLUMNS:
        raise ValueError("عدد أعمدة ملف Excel أكبر من الحد المسموح")
    if sheet.max_row and sheet.max_row > MAX_WORKBOOK_ROWS + 1:
        raise ValueError("عدد صفوف ملف Excel أكبر من الحد المسموح")


def normalize_pharmacy_status(value: Any) -> str:
    raw = str(value or "active").strip().lower()
    try:
        return _STATUS_ALIASES[raw]
    except KeyError as exc:
        raise ValueError(f"حالة الصيدلية غير صالحة: {value}") from exc


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower()


def _header_map(values: list[Any], definitions: dict[str, set[str]]) -> dict[str, int]:
    normalized = [_normalize_header(value) for value in values]
    result: dict[str, int] = {}
    for key, aliases in definitions.items():
        for index, value in enumerate(normalized):
            if value in aliases:
                result[key] = index
                break
    return result


def parse_shifts_workbook(data: bytes) -> list[ParsedShift]:
    _validate_xlsx_archive(data)
    try:
        workbook = load_workbook(BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError("تعذر فتح ملف Excel؛ تأكد أنه ملف xlsx صالح") from exc
    sheet = workbook.active
    _validate_sheet_size(sheet)
    rows = sheet.iter_rows(values_only=True)
    try:
        header = list(next(rows))
    except StopIteration as exc:
        raise ValueError("ملف Excel فارغ") from exc
    mapping = _header_map(header, SHIFT_HEADERS)
    missing = [key for key in ("pharmacy", "date", "start", "end") if key not in mapping]
    if missing:
        raise ValueError("أعمدة ملف المناوبات غير مكتملة")

    parsed: list[ParsedShift] = []
    for row_number, values in enumerate(rows, start=2):
        values = list(values)
        if not any(value not in (None, "") for value in values):
            continue
        try:
            pharmacy_name = str(values[mapping["pharmacy"]] or "").strip()
            duty_date = parse_date_value(_excel_date_value(values[mapping["date"]]))
            start_time = parse_time_value(_excel_time_value(values[mapping["start"]]))
            end_time = parse_time_value(_excel_time_value(values[mapping["end"]]))
        except Exception as exc:
            raise ValueError(f"خطأ في السطر {row_number}: {exc}") from exc
        parsed.append(
            ParsedShift(
                pharmacy_name=pharmacy_name,
                duty_date=duty_date,
                start_time=start_time,
                end_time=end_time,
                row_number=row_number,
                raw_data={
                    "pharmacy": pharmacy_name,
                    "date": str(values[mapping["date"]]),
                    "start": str(values[mapping["start"]]),
                    "end": str(values[mapping["end"]]),
                },
            )
        )
    if not parsed:
        raise ValueError("لم يتم العثور على مناوبات داخل الملف")
    return parsed


def parse_pharmacies_workbook(data: bytes) -> list[dict[str, Any]]:
    _validate_xlsx_archive(data)
    try:
        workbook = load_workbook(BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError("تعذر فتح ملف Excel") from exc
    sheet = workbook.active
    _validate_sheet_size(sheet)
    rows = sheet.iter_rows(values_only=True)
    try:
        header = list(next(rows))
    except StopIteration as exc:
        raise ValueError("ملف Excel فارغ") from exc
    mapping = _header_map(header, PHARMACY_HEADERS)
    if "name" not in mapping or "address" not in mapping:
        raise ValueError("يجب أن يحتوي الملف على اسم الصيدلية والعنوان")
    result: list[dict[str, Any]] = []
    for row_number, values in enumerate(rows, start=2):
        values = list(values)
        if not any(value not in (None, "") for value in values):
            continue
        name = str(values[mapping["name"]] or "").strip()
        address = str(values[mapping["address"]] or "").strip()
        aliases_raw = str(values[mapping["aliases"]] or "") if "aliases" in mapping else ""
        aliases = [item.strip() for item in aliases_raw.replace("،", ",").split(",") if item.strip()]
        status = (
            normalize_pharmacy_status(values[mapping["status"]])
            if "status" in mapping
            else "active"
        )
        notes = str(values[mapping["notes"]] or "").strip() if "notes" in mapping else ""
        if not name or not address:
            raise ValueError(f"السطر {row_number}: الاسم أو العنوان فارغ")
        result.append(
            {
                "row_number": row_number,
                "name": name,
                "address": address,
                "aliases": aliases,
                "status": status,
                "notes": notes or None,
            }
        )
    return result


def build_shifts_template() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "المناوبات"
    headers = ["اسم الصيدلية", "التاريخ", "وقت البداية", "وقت النهاية"]
    sheet.append(headers)
    sheet.append(["صيدلية الشفاء", date.today(), "8:00 PM", "8:00 AM"])
    _style_sheet(sheet, widths=[28, 18, 18, 18])
    sheet.freeze_panes = "A2"
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_pharmacies_template() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "الصيدليات"
    headers = ["اسم الصيدلية", "العنوان", "الأسماء البديلة", "الحالة", "ملاحظات"]
    sheet.append(headers)
    sheet.append(["صيدلية الشفاء", "شارع البلدية", "الشفاء، الشفا", "active", ""])
    _style_sheet(sheet, widths=[28, 45, 35, 16, 30])
    sheet.freeze_panes = "A2"
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_missing_pharmacies_template(names: Iterable[str], batch_id: int) -> bytes:
    """Create a pre-filled workbook for all unmatched pharmacy names in a draft."""
    unique_names = sorted({name.strip() for name in names if name and name.strip()})
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "الصيدليات الناقصة"
    sheet.append(["اسم الصيدلية", "العنوان", "الأسماء البديلة", "الحالة", "ملاحظات"])
    for name in unique_names:
        sheet.append([name, "", "", "active", f"أضيفت من مسودة المناوبات #{batch_id}"])
    _style_sheet(sheet, widths=[32, 50, 35, 16, 36])
    sheet.freeze_panes = "A2"
    address_fill = PatternFill("solid", fgColor="FFF2CC")
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row=row, column=2).fill = address_fill
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_pharmacies(pharmacies: list[Any]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "الصيدليات"
    sheet.append(["المعرّف", "اسم الصيدلية", "العنوان", "الأسماء البديلة", "الحالة", "ملاحظات"])
    for pharmacy in pharmacies:
        sheet.append(
            [
                pharmacy.id,
                pharmacy.name,
                pharmacy.address,
                "، ".join(alias.alias for alias in pharmacy.aliases),
                pharmacy.status,
                pharmacy.notes or "",
            ]
        )
    _style_sheet(sheet, widths=[12, 30, 50, 40, 18, 35])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_shifts(shifts: list[Any], timezone) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "المناوبات"
    sheet.append(["المعرّف", "اسم الصيدلية", "التاريخ", "وقت البداية", "وقت النهاية"])
    for shift in shifts:
        start = shift.start_at.astimezone(timezone)
        end = shift.end_at.astimezone(timezone)
        sheet.append(
            [
                shift.id,
                shift.pharmacy.name,
                start.date(),
                start.strftime("%I:%M %p"),
                end.strftime("%I:%M %p"),
            ]
        )
    _style_sheet(sheet, widths=[12, 30, 18, 18, 18])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _excel_date_value(value: Any) -> str | date | datetime:
    if isinstance(value, (datetime, date)):
        return value
    return str(value or "")


def _excel_time_value(value: Any) -> str | time | datetime:
    if isinstance(value, (datetime, time)):
        return value
    if isinstance(value, (int, float)) and 0 <= value < 1:
        seconds = round(value * 24 * 60 * 60)
        hours, remainder = divmod(seconds, 3600)
        minutes = remainder // 60
        return time(hours % 24, minutes)
    return str(value or "")


def _style_sheet(sheet, widths: list[int]) -> None:
    fill = PatternFill("solid", fgColor="198754")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.sheet_view.rightToLeft = True
